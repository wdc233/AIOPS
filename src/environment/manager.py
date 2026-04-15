"""Environment manager for cluster and server information."""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Dict, List, Optional, Union

from src.db.repository import ClusterRepository
from src.models.types import ClusterInfo, ServerInfo

logger = logging.getLogger(__name__)


class EnvironmentManager:
    """Manager for global cluster and server information."""

    def __init__(self) -> None:
        """Initialize environment manager."""
        self._cluster_repo = ClusterRepository()
        self._clusters: Dict[str, ClusterInfo] = {}
        self._servers: Dict[str, ServerInfo] = {}  # ip -> ServerInfo

    async def initialize(self, config_path: Optional[Union[str, Path]] = None) -> None:
        """Initialize environment from config file or database.

        Args:
            config_path: Path to JSON config file with cluster/server info
        """
        if config_path:
            try:
                await self._load_from_file(str(config_path))
                return
            except FileNotFoundError:
                pass
        await self._load_from_database()

    async def _load_from_file(self, config_path: str) -> None:
        """Load environment from JSON or Excel config file."""
        path = Path(config_path)
        if path.suffix == ".xlsx":
            await self._load_from_xlsx(path)
        else:
            await self._load_from_json(path)

    async def _load_from_json(self, path: Path) -> None:
        """Load environment from JSON config file."""
        with open(path, "r", encoding="utf-8") as f:
            config = json.load(f)

        clusters_data = config.get("clusters", [])
        for cluster_data in clusters_data:
            cluster = ClusterInfo.model_validate(cluster_data)
            self._clusters[cluster.cluster_name] = cluster
            for server in cluster.servers:
                self._servers[server.ip] = server

        logger.info(f"Loaded {len(self._clusters)} clusters from JSON file")

    async def _load_from_xlsx(self, path: Path) -> None:
        """Load environment from Excel config file."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required to load Excel files. Install with: pip install openpyxl")

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        required_cols = {"cluster_name", "cluster_type", "env", "ip", "username"}
        missing = required_cols - set(headers)
        if missing:
            raise ValueError(f"Excel file missing required columns: {missing}")

        cluster_servers: Dict[str, List[Dict]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data = dict(zip(headers, row))
            cluster_name = data["cluster_name"]
            if cluster_name not in cluster_servers:
                cluster_servers[cluster_name] = []
            cluster_servers[cluster_name].append(data)

        for cluster_name, servers_data in cluster_servers.items():
            first = servers_data[0]
            cluster = ClusterInfo(
                cluster_name=cluster_name,
                cluster_type=first["cluster_type"],
                env=first["env"],
                servers=[
                    ServerInfo(
                        ip=srv["ip"],
                        port=srv.get("port", 22),
                        username=srv["username"],
                        password=srv.get("password"),
                        cluster_name=cluster_name,
                    )
                    for srv in servers_data
                ],
            )
            self._clusters[cluster.cluster_name] = cluster
            for server in cluster.servers:
                self._servers[server.ip] = server

        logger.info(f"Loaded {len(self._clusters)} clusters from Excel file")

    async def _load_from_database(self) -> None:
        """Load environment from database."""
        clusters = await self._cluster_repo.list_clusters()
        for cluster in clusters:
            self._clusters[cluster.cluster_name] = cluster
            for server in cluster.servers:
                self._servers[server.ip] = server

        logger.info(f"Loaded {len(self._clusters)} clusters from database")

    def get_cluster(self, cluster_name: str) -> Optional[ClusterInfo]:
        """Get cluster by name."""
        return self._clusters.get(cluster_name)

    def get_all_clusters(self) -> List[ClusterInfo]:
        """Get all clusters."""
        return list(self._clusters.values())

    def get_server(self, ip: str) -> Optional[ServerInfo]:
        """Get server by IP."""
        return self._servers.get(ip)

    def get_servers_by_cluster(self, cluster_name: str) -> List[ServerInfo]:
        """Get all servers in a cluster."""
        cluster = self._clusters.get(cluster_name)
        return cluster.servers if cluster else []

    def add_cluster(self, cluster: ClusterInfo) -> None:
        """Add or update a cluster."""
        self._clusters[cluster.cluster_name] = cluster
        for server in cluster.servers:
            self._servers[server.ip] = server

    def add_server(self, server: ServerInfo) -> None:
        """Add or update a server."""
        self._servers[server.ip] = server
        # Add to cluster if exists
        cluster = self._clusters.get(server.cluster_name)
        if cluster:
            for i, s in enumerate(cluster.servers):
                if s.ip == server.ip:
                    cluster.servers[i] = server
                    return
            cluster.servers.append(server)

    def remove_cluster(self, cluster_name: str) -> None:
        """Remove a cluster and its servers."""
        cluster = self._clusters.pop(cluster_name, None)
        if cluster:
            for server in cluster.servers:
                self._servers.pop(server.ip, None)

    def remove_server(self, ip: str) -> None:
        """Remove a server."""
        server = self._servers.pop(ip, None)
        if server:
            cluster = self._clusters.get(server.cluster_name)
            if cluster:
                cluster.servers = [s for s in cluster.servers if s.ip != ip]

    async def save_to_database(self) -> None:
        """Save current environment to database."""
        for cluster in self._clusters.values():
            try:
                await self._cluster_repo.create_cluster(cluster)
            except Exception:
                pass  # Cluster might already exist

            for server in cluster.servers:
                try:
                    await self._cluster_repo.add_server(server)
                except Exception:
                    pass  # Server might already exist

        logger.info(f"Saved {len(self._clusters)} clusters to database")


# Global environment manager instance
_env_manager: Optional[EnvironmentManager] = None


def get_environment_manager() -> EnvironmentManager:
    """Get global environment manager instance."""
    global _env_manager
    if _env_manager is None:
        _env_manager = EnvironmentManager()
    return _env_manager


async def initialize_environment(config_path: Optional[str] = None) -> EnvironmentManager:
    """Initialize and get environment manager."""
    manager = get_environment_manager()
    await manager.initialize(config_path)
    return manager